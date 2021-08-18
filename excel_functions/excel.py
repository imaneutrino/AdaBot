from openpyxl import load_workbook
from helpers.convertUnityStringToInteger import get_types_through_key


def bloco_information(filename):
    # Available fields to register bloco
    bloco_data = {
        "titulo": "",
        "entrega": "",
        "more_info": "",
        "display_order": 1
    }
    wb = load_workbook(filename=filename)
    sheet = wb.active
    bloco_data["titulo"] = sheet['B2'].value
    bloco_data["entrega"] = sheet['C2'].value
    bloco_data["more_info"] = sheet['D2'].value
    return bloco_data


def tipo_information(filename):
    # Available fields to register tipo
    tipo_data = {
        "titulo": "",
        "categoria": "",
        "display_order": 0
    }
    wb = load_workbook(filename=filename)
    sheet = wb.active
    tipo_data["titulo"] = sheet["E2"].value
    tipo_data["categoria"] = get_types_through_key(sheet["F2"].value)
    tipo_data["display_order"] = sheet["G2"].value
    return tipo_data
